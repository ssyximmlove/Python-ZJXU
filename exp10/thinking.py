import random

import openpyxl


def generate_mock_student_data(num_students=6):
    names = ['Alice', 'Bob', 'Charlie', 'David', 'Emily', 'Frank']
    genders = ['Male', 'Female']
    classes = ['Class A', 'Class B', 'Class C']
    data = []
    for i in range(num_students):
        name = random.choice(names)
        gender = random.choice(genders)
        classroom = random.choice(classes)
        student_id = 2024000 + i
        data.append([name, gender, classroom, student_id])
    return data


def write_to_excel(filename, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['姓名', '性别', '班级', '学号'])
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def read_and_print_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            print(row)
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")


if __name__ == "__main__":
    filename = 'student_info.xlsx'
    mock_data = generate_mock_student_data()
    write_to_excel(filename, mock_data)
    read_and_print_excel(filename)
