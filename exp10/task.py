import random

import openpyxl


def generate_mock_data(num_rows=20):
    names = ['Alice', 'Bob', 'Charlie', 'David', 'Emily']
    courses = ['Math', 'English', 'Science', 'History', 'Art']
    data = []
    for _ in range(num_rows):
        name = random.choice(names)
        course = random.choice(courses)
        score = random.randint(50, 100)
        data.append([name, course, score])
    return data


def write_to_excel(filename, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['姓名', '课程', '成绩'])
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def calculate_highest_scores(input_filename, output_filename):
    try:
        workbook = openpyxl.load_workbook(input_filename)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
    except FileNotFoundError:
        print(f"Error: The file '{input_filename}' was not found.")
        return

    highest_scores = {}
    for row in data:
        name, course, score = row
        if (name, course) not in highest_scores:
            highest_scores[(name, course)] = score
        else:
            highest_scores[(name, course)] = max(highest_scores[(name, course)], score)

    output_data = [['姓名', '课程', '成绩']]
    for (name, course), score in highest_scores.items():
        output_data.append([name, course, score])

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in output_data:
        sheet.append(row)
    workbook.save(output_filename)
    print(f"Highest scores calculated and saved to '{output_filename}'")


if __name__ == "__main__":
    input_filename = 'student_scores.xlsx'
    output_filename = 'highest_scores.xlsx'
    mock_data = generate_mock_data()
    write_to_excel(input_filename, mock_data)

    calculate_highest_scores(input_filename, output_filename)
